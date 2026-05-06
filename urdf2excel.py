#!/usr/bin/env python3
"""
urdf2excel.py  ·  URDF joint limits & link inertia → Excel

Usage:
  python urdf2excel.py path/to/robot.urdf              # single file (recommended)
  python urdf2excel.py path/to/robot.urdf -o out.xlsx  # custom output name
  python urdf2excel.py --robots robots/                # scan whole directory
  python urdf2excel.py --list                          # list found URDF files only
"""

from __future__ import annotations

import argparse
import math
import sys
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] openpyxl missing.  Run: pip install openpyxl")

# ── paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR         = Path(__file__).parent
DEFAULT_ROBOTS_DIR = SCRIPT_DIR / "robots"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "output"

# ── style helpers ─────────────────────────────────────────────────────────────
_SIDE   = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_ROW_FILLS = [_fill("EFF3F8"), _fill("FFFFFF")]

# ── URDF parsing ──────────────────────────────────────────────────────────────
def _to_float(s: str | None) -> float | None:
    try:
        return float(s) if s is not None else None
    except (ValueError, TypeError):
        return None

# Joint types that carry limits (skip fixed / floating / planar / mimic-only)
_MOVABLE_TYPES = {"revolute", "prismatic", "continuous"}

def parse_urdf(urdf_path: Path) -> tuple[str, list[dict], list[dict]]:
    """Return (robot_name, joints, links_with_inertia)."""
    tree = ET.parse(urdf_path)
    root = tree.getroot()
    robot_name = root.get("name", urdf_path.stem)

    # ── joints ────────────────────────────────────────────────────────────────
    joints: list[dict] = []
    for j in root.findall("joint"):
        jtype = j.get("type", "fixed")
        limit  = j.find("limit")
        # include if movable OR has limit element
        if jtype not in _MOVABLE_TYPES and limit is None:
            continue

        parent = j.find("parent")
        child  = j.find("child")
        axis   = j.find("axis")
        origin = j.find("origin")
        dyn    = j.find("dynamics")

        lower  = _to_float(limit.get("lower"))   if limit is not None else None
        upper  = _to_float(limit.get("upper"))   if limit is not None else None
        effort = _to_float(limit.get("effort"))  if limit is not None else None
        vel    = _to_float(limit.get("velocity")) if limit is not None else None

        joints.append({
            "name":     j.get("name", ""),
            "type":     jtype,
            "parent":   parent.get("link", "") if parent is not None else "",
            "child":    child.get("link",  "") if child  is not None else "",
            "axis":     axis.get("xyz", "") if axis is not None else "",
            "orig_xyz": origin.get("xyz", "") if origin is not None else "",
            "orig_rpy": origin.get("rpy", "") if origin is not None else "",
            "lower":    lower,
            "upper":    upper,
            "effort":   effort,
            "velocity": vel,
            "damping":  _to_float(dyn.get("damping"))  if dyn is not None else None,
            "friction": _to_float(dyn.get("friction")) if dyn is not None else None,
        })

    # ── links (inertial) ──────────────────────────────────────────────────────
    links: list[dict] = []
    for lk in root.findall("link"):
        inertial = lk.find("inertial")
        if inertial is None:
            continue

        mass_el    = inertial.find("mass")
        origin     = inertial.find("origin")
        inertia_el = inertial.find("inertia")

        com_x = com_y = com_z = None
        com_rpy = ""
        if origin is not None:
            xyz = origin.get("xyz", "0 0 0").split()
            com_x   = _to_float(xyz[0]) if len(xyz) > 0 else None
            com_y   = _to_float(xyz[1]) if len(xyz) > 1 else None
            com_z   = _to_float(xyz[2]) if len(xyz) > 2 else None
            com_rpy = origin.get("rpy", "")

        ivals: dict[str, float | None] = {
            k: (_to_float(inertia_el.get(k)) if inertia_el is not None else None)
            for k in ("ixx", "ixy", "ixz", "iyy", "iyz", "izz")
        }

        links.append({
            "name":    lk.get("name", ""),
            "mass":    _to_float(mass_el.get("value")) if mass_el is not None else None,
            "com_x":   com_x,
            "com_y":   com_y,
            "com_z":   com_z,
            "com_rpy": com_rpy,
            **ivals,
        })

    return robot_name, joints, links


def find_urdfs(robots_dir: Path) -> list[Path]:
    return sorted(robots_dir.rglob("*.urdf"))


# ── Excel: joint limits sheet ─────────────────────────────────────────────────
#  Row 1  │ 关节名称 │ ←── 弧度值 (rad) ──→ │ ←── 角度值 (°) ──→ │ 力矩 │ 速度 │ 类型 │
#  Row 2  │          │  最大值  │  最小值  │  最大值  │  最小值  │      │      │      │
#  Row 3+ │  data…   │

_COL_JOINT  = 1   # A
_COL_R_MAX  = 2   # B  ┐ merged B1:C1 "弧度值 (rad)"
_COL_R_MIN  = 3   # C  ┘
_COL_D_MAX  = 4   # D  ┐ merged D1:E1 "角度值 (°)"
_COL_D_MIN  = 5   # E  ┘
_COL_EFFORT = 6   # F
_COL_VEL    = 7   # G
_COL_TYPE   = 8   # H


def _cell(ws, row: int, col: int, value: Any = None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = _BORDER
    return c


def write_joint_sheet(ws, joints: list[dict], robot_name: str) -> None:
    ws.title = "关节限位"

    # ── Row 1: group headers ──────────────────────────────────────────────────
    r1_specs = [
        (_COL_JOINT,  "关节名称",      "1C3553", 1),
        (_COL_R_MAX,  "弧度值 (rad)",  "2E6096", 2),   # spans B-C
        (_COL_D_MAX,  "角度值 (°)",    "2E6096", 2),   # spans D-E
        (_COL_EFFORT, "力矩 (N·m)",    "1C3553", 1),
        (_COL_VEL,    "速度 (rad/s)",  "1C3553", 1),
        (_COL_TYPE,   "类型",          "1C3553", 1),
    ]
    for col, label, bg, span in r1_specs:
        c = _cell(ws, 1, col, label)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        if span == 2:
            end_col = get_column_letter(col + 1)
            ws.merge_cells(f"{get_column_letter(col)}1:{end_col}1")

    # ── Row 2: sub-headers ────────────────────────────────────────────────────
    r2_specs = [
        (_COL_JOINT,  ""),
        (_COL_R_MAX,  "最大值"),
        (_COL_R_MIN,  "最小值"),
        (_COL_D_MAX,  "最大值"),
        (_COL_D_MIN,  "最小值"),
        (_COL_EFFORT, ""),
        (_COL_VEL,    ""),
        (_COL_TYPE,   ""),
    ]
    for col, label in r2_specs:
        c = _cell(ws, 2, col, label)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill("4A7FB5")
        c.alignment = _align("center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    def _r(v: float | None, digits: int = 6) -> float | str:
        return round(v, digits) if v is not None else ""

    def _deg(v: float | None) -> float | str:
        return round(math.degrees(v), 2) if v is not None else ""

    for row_i, j in enumerate(joints, 3):
        fill = _ROW_FILLS[row_i % 2]
        data = [
            (1, j["name"],         "left"),
            (2, _r(j["upper"]),    "right"),
            (3, _r(j["lower"]),    "right"),
            (4, _deg(j["upper"]),  "right"),
            (5, _deg(j["lower"]),  "right"),
            (6, _r(j["effort"], 2) if j["effort"] is not None else "", "right"),
            (7, _r(j["velocity"], 2) if j["velocity"] is not None else "", "right"),
            (8, j["type"],         "center"),
        ]
        for col, val, halign in data:
            c = _cell(ws, row_i, col, val)
            c.fill      = fill
            c.alignment = _align(halign, "center")

    # column widths
    widths = {1: 32, 2: 11, 3: 11, 4: 11, 5: 11, 6: 13, 7: 14, 8: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"


# ── Excel: link inertia sheet ─────────────────────────────────────────────────

def write_inertia_sheet(ws, links: list[dict]) -> None:
    ws.title = "连杆惯量"

    headers = [
        ("连杆名称",      1),
        ("质量 (kg)",     2),
        ("重心 X (m)",    3),
        ("重心 Y (m)",    4),
        ("重心 Z (m)",    5),
        ("Ixx (kg·m²)",   6),
        ("Ixy (kg·m²)",   7),
        ("Ixz (kg·m²)",   8),
        ("Iyy (kg·m²)",   9),
        ("Iyz (kg·m²)",  10),
        ("Izz (kg·m²)",  11),
    ]
    keys = ["name", "mass", "com_x", "com_y", "com_z",
            "ixx", "ixy", "ixz", "iyy", "iyz", "izz"]

    for label, col in headers:
        c = _cell(ws, 1, col, label)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill("1C3553")
        c.alignment = _align("center", wrap=True)

    ws.row_dimensions[1].height = 36

    for row_i, lk in enumerate(links, 2):
        fill = _ROW_FILLS[row_i % 2]
        for col, key in enumerate(keys, 1):
            val = lk.get(key)
            if val is None:
                val = ""
            c = _cell(ws, row_i, col, val)
            c.fill      = fill
            c.alignment = _align("right" if col > 1 else "left", "center")
            # plain decimal format (up to 10 decimal places, no trailing zeros)
            if isinstance(val, float):
                if col == 2:                         # mass: 6 dp
                    c.number_format = "0.000000"
                elif col in (3, 4, 5):               # CoM: 6 dp
                    c.number_format = "0.000000"
                else:                                # inertia: 10 dp
                    c.number_format = "0.0000000000"

    # column widths
    widths = {1: 32, 2: 12, 3: 14, 4: 14, 5: 14,
              6: 18, 7: 18, 8: 18, 9: 18, 10: 18, 11: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── top-level write ───────────────────────────────────────────────────────────

def write_excel(
    output_path: Path,
    robot_name: str,
    joints: list[dict],
    links: list[dict],
) -> None:
    wb = openpyxl.Workbook()

    ws_j = wb.active
    write_joint_sheet(ws_j, joints, robot_name)

    ws_i = wb.create_sheet()
    write_inertia_sheet(ws_i, links)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract URDF joint limits & link inertia → Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "urdf", nargs="?", default=None,
        help="Path to a single URDF file  (recommended)",
    )
    parser.add_argument(
        "-o", "--output", default=None,
        help="Output .xlsx path  (default: output/<robot>_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--robots", default=None,
        help=f"Scan a robots directory recursively  (default: {DEFAULT_ROBOTS_DIR})",
    )
    parser.add_argument(
        "--list", action="store_true",
        help="List found URDF files and exit",
    )
    args = parser.parse_args()

    # ── determine URDF files to process ───────────────────────────────────────
    if args.urdf:
        urdf_path = Path(args.urdf).resolve()
        if not urdf_path.exists():
            sys.exit(f"[ERROR] File not found: {urdf_path}")
        urdf_files = [urdf_path]
    else:
        robots_dir = Path(args.robots) if args.robots else DEFAULT_ROBOTS_DIR
        if not robots_dir.exists():
            sys.exit(f"[ERROR] Directory not found: {robots_dir}")
        urdf_files = find_urdfs(robots_dir)
        if not urdf_files:
            sys.exit(f"[ERROR] No .urdf files found under {robots_dir}")

    if args.list:
        print(f"\nFound {len(urdf_files)} URDF file(s):")
        for u in urdf_files:
            print(f"  {u}")
        return

    # ── process each file ─────────────────────────────────────────────────────
    for urdf_path in urdf_files:
        print(f"\nParsing: {urdf_path}")
        try:
            robot_name, joints, links = parse_urdf(urdf_path)
        except ET.ParseError as e:
            print(f"  [WARN] Parse error: {e}")
            continue

        print(f"  {len(joints)} movable joints,  {len(links)} links w/ inertia")

        if args.output and len(urdf_files) == 1:
            out = Path(args.output)
        else:
            ts  = datetime.now().strftime("%Y%m%d_%H%M")
            out = DEFAULT_OUTPUT_DIR / f"{urdf_path.stem}_{ts}.xlsx"

        write_excel(out, robot_name, joints, links)
        print(f"  ✓ Saved → {out.resolve()}")

    print()


if __name__ == "__main__":
    main()

